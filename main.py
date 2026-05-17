from openpyxl import load_workbook
import pandas as pd
import numpy as np
import re

import requests as req
import datetime as dt
import locale

TEMPLATE = 'template.xlsx'
TEMPLATE_SHEET = 'Protokoll'

START_ROW = 'Stadtratssitzung'

SITZUNGEN = 'https://stadtrat.bern.ch/de/sitzungen/'
PARSE_LINK = 'Stadtrat'


DATE_TO_REPLACE = 'DATE_FROM'
DATA_LINK = f'https://stadtrat.bern.ch/format/module/politik_axioma/sitzungen/sitzungen_data_server.php?dateFrom={DATE_TO_REPLACE}&dateTo='

TRAKTANDUM_NUMBER_LABEL = "Traktandumsnummer"
TRAKTANDEN_LABEL = 'Traktanden'
TRAKTANDUM_LABEL = 'Traktandum'
TITEL_LABEL = 'Titel'
GESCHAEFTSNR_LABEL = 'Geschaeftsnummer'
LOCALE = 'de_CH.UTF-8'


ORT_LABEL = 'Berner Rathaus, Rathausplatz 2, 3011 Bern, Grossratssaal'

def main():
    locale.setlocale(locale.LC_TIME, LOCALE)

    template = pd.read_excel(TEMPLATE)
    start_row = np.argwhere(template.iloc[:, 0] == START_ROW).flatten()[0] + 2
    date_row = start_row + 1
    start_traktanden = start_row + 2

    sitzungen = load_sitzungen()['data']

    stadtrats_sitzungen = [entry for entry in sitzungen if re.match(f".*{PARSE_LINK}.*", entry[TITEL_LABEL])]
    print(f"Found {len(stadtrats_sitzungen)} {PARSE_LINK} entries")

    if len(stadtrats_sitzungen) == 0:
        return;


    next_sitzung = stadtrats_sitzungen[0]
    date_str = next_sitzung['Datum']['Text']
    date = dt.datetime.strptime(date_str, '%d.%m.%Y').strftime('%A, %d. %B %Y')
    times = [time['Von'] + "-" + time['Bis'] + ' Uhr' for time in next_sitzung['Sitzungsdaten']['Sitzungsdatum']]

    datetimestring = date + ", " + ' und '.join(times)

    traktanden = [(trakt[TRAKTANDUM_NUMBER_LABEL], get_string_for_traktandum(trakt)) for trakt in next_sitzung[TRAKTANDEN_LABEL][TRAKTANDUM_LABEL]]

    wb = load_workbook(TEMPLATE)
    ws = wb[TEMPLATE_SHEET]

    ws.cell(row=date_row, column=1, value=f"{datetimestring}\n{ORT_LABEL}")
    for (i, trakt) in enumerate(traktanden):
        ws.cell(row=start_traktanden+i, column=3, value=trakt[0])
        ws.cell(row=start_traktanden+i, column=4, value=trakt[1])

    filename = f'Protokoll Fraktionssitzung {date_str}.xlsx'
    print(f'Writing to "{filename}"')
    wb.save(filename)

def get_string_for_traktandum(traktandum):
    title = traktandum[TITEL_LABEL]
    number = traktandum[GESCHAEFTSNR_LABEL]
    return title + "\n" + number if isinstance(number, str) else title


def load_sitzungen(dateFrom=''):

    """Loads sitzungen starting from the given date or today

    Parameters
    ----------
    dateFrom : string
        Date of the form 'dd.mm.yyyy'

    Returns
    -------
    Array
        JSON array containing the server response

    """
    """Loads sitzungen starting from date given or today

    Parameters
    ----------
    dateFrom : string containing a date of the form DD.MM.YYYY

    """
    if dateFrom == '':
        dateFrom = dt.datetime.now().strftime('%d.%m.%Y')

    url = DATA_LINK.replace(DATE_TO_REPLACE, dateFrom)

    return req.get(url).json()


if __name__ == "__main__":
    main()
